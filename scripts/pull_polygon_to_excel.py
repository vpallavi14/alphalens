"""
Pull Polygon.io daily aggregate data for multiple tickers and save to Excel.
Run: python scripts/pull_polygon_to_excel.py
Output: scripts/output/polygon_stock_data.xlsx
"""

import os, time, requests
from datetime import date, timedelta
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

API_KEY  = os.environ.get("MASSIVE_API_KEY", "0pWckqzeoG1dxGtfu_pZrpPL_JKoQbvk")
BASE_URL = "https://api.polygon.io"

TICKERS   = ["AAPL", "MSFT", "NVDA", "TSLA", "GOOGL", "AMZN", "META", "NFLX", "AMD", "INTC"]
END_DATE  = date.today() - timedelta(days=1)
START_DATE = END_DATE - timedelta(days=30)   # last ~30 calendar days (~20 trading days)

OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "polygon_stock_data.xlsx")

# ── Colours ──────────────────────────────────────────────────
HEADER_BG  = "0F172A"   # dark slate
HEADER_FG  = "F1F5F9"
UP_COLOR   = "064E3B"   # dark green bg
DOWN_COLOR = "450A0A"   # dark red bg
ALT_ROW    = "1E293B"

def ts_ms_to_date(ms: int) -> str:
    from datetime import datetime
    return datetime.utcfromtimestamp(ms / 1000).strftime("%Y-%m-%d")

def ts_ms_to_datetime(ms: int) -> str:
    from datetime import datetime
    return datetime.utcfromtimestamp(ms / 1000).strftime("%Y-%m-%d %H:%M:%S")

def fetch_aggs(ticker: str) -> list:
    url = f"{BASE_URL}/v2/aggs/ticker/{ticker}/range/1/day/{START_DATE}/{END_DATE}"
    r = requests.get(url, params={
        "apiKey": API_KEY, "adjusted": "true", "sort": "asc", "limit": 50
    }, timeout=15)
    data = r.json()
    if data.get("status") == "OK" and data.get("results"):
        return data["results"]
    print(f"  [WARN] {ticker}: {data.get('status')} — {data.get('error', 'no results')}")
    return []

def style_header(cell, bold=True):
    cell.font      = Font(name="Arial", bold=bold, color=HEADER_FG, size=10)
    cell.fill      = PatternFill("solid", start_color=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def thin_border():
    s = Side(style="thin", color="334155")
    return Border(left=s, right=s, top=s, bottom=s)

def build_ticker_sheet(wb: openpyxl.Workbook, ticker: str, bars: list):
    ws = wb.create_sheet(ticker)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{ticker} — Daily Bars (Polygon.io)  |  {START_DATE} → {END_DATE}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=12, color="10B981")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Headers — added Unix ms + full UTC datetime columns
    headers = ["Unix Timestamp (ms)", "Date (UTC)", "Datetime (UTC)", "Open", "High", "Low", "Close", "Volume", "VWAP", "Trades", "Day Δ $", "Day Δ %"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        style_header(cell)
        cell.border = thin_border()
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A1:L1")  # extend title across all 12 cols

    # Data rows
    for row_idx, bar in enumerate(bars, 3):
        bar_date     = ts_ms_to_date(bar["t"])
        bar_datetime = ts_ms_to_datetime(bar["t"]) + " UTC"
        prev_close   = bars[row_idx - 4]["c"] if row_idx > 3 else bar["c"]
        delta        = bar["c"] - prev_close
        delta_pct    = (delta / prev_close) if prev_close else 0  # decimal for Excel % format
        is_up        = delta >= 0

        row_fill = PatternFill("solid", start_color=ALT_ROW if row_idx % 2 == 0 else "0F172A")

        values = [
            bar["t"],           # raw Unix ms
            bar_date,           # YYYY-MM-DD
            bar_datetime,       # YYYY-MM-DD HH:MM:SS UTC
            bar["o"], bar["h"], bar["l"], bar["c"],
            bar["v"], bar.get("vw", 0), bar.get("n", 0),
            round(delta, 2) if row_idx > 3 else None,
            round(delta_pct, 2) if row_idx > 3 else None
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10, color="CBD5E1")
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal="right" if col_idx != 3 else "left")
            cell.border    = thin_border()

            # Col 1 = Unix ms, Col 2 = Date, Col 3 = Datetime, Col 4-7 = OHLC
            # Col 8 = Volume, Col 9 = VWAP, Col 10 = Trades, Col 11 = Δ$, Col 12 = Δ%

            if col_idx == 1:   # Unix ms — raw number
                cell.number_format = '#,##0'
                cell.font = Font(name="Arial", size=9, color="475569")
            elif col_idx in (2, 3):  # date / datetime strings
                cell.alignment = Alignment(horizontal="left")
            elif col_idx == 7:  # Close — colour green/red
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color="34D399" if is_up else "F87171")
                cell.number_format = '$#,##0.00'
            elif col_idx in (4, 5, 6, 9):  # Open, High, Low, VWAP
                cell.number_format = '$#,##0.00'
            elif col_idx in (8, 10):  # Volume, Trades
                cell.number_format = '#,##0'
            elif col_idx == 11 and val is not None:  # Δ$
                cell.number_format = '$#,##0.00;($#,##0.00);"-"'
                cell.font = Font(name="Arial", size=10,
                                 color="34D399" if val >= 0 else "F87171")
            elif col_idx == 12 and val is not None:  # Δ%
                cell.number_format = '0.00%;-0.00%;"-"'
                cell.font = Font(name="Arial", size=10,
                                 color="34D399" if val >= 0 else "F87171")

    # Column widths: Unix ms, Date, Datetime, O, H, L, C, Vol, VWAP, Trades, Δ$, Δ%
    widths = [18, 13, 22, 10, 10, 10, 10, 14, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Source note
    note_row = len(bars) + 4
    ws.cell(row=note_row, column=1, value=f"Source: Polygon.io /v2/aggs · Adjusted=True · Pulled {date.today()}")
    ws.cell(row=note_row, column=1).font = Font(name="Arial", size=9, color="475569", italic=True)

def build_summary_sheet(wb: openpyxl.Workbook, all_data: dict):
    ws = wb.active
    ws.title   = "Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = f"AlphaLens — Polygon Stock Summary  |  Latest close as of {END_DATE}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="10B981")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Ticker", "Latest Date", "Close ($)", "Open ($)", "High ($)", "Low ($)", "Volume", "30D Δ %"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        style_header(cell)
        cell.border = thin_border()
    ws.row_dimensions[2].height = 18

    for row_idx, (ticker, bars) in enumerate(all_data.items(), 3):
        if not bars:
            continue
        latest = bars[-1]
        first  = bars[0]
        # Store as decimal (0.05 = 5%) so Excel % format displays correctly
        change_30d = ((latest["c"] - first["c"]) / first["c"]) if first["c"] else 0
        is_up = change_30d >= 0
        alt   = PatternFill("solid", start_color=ALT_ROW if row_idx % 2 == 0 else "0F172A")

        row_data = [
            ticker, ts_ms_to_date(latest["t"]), latest["c"],
            latest["o"], latest["h"], latest["l"], latest["v"], round(change_30d, 4)
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10, color="CBD5E1")
            cell.fill      = alt
            cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "center")
            cell.border    = thin_border()

            if col_idx in (3, 4, 5, 6):
                cell.number_format = '$#,##0.00'
            elif col_idx == 7:
                cell.number_format = '#,##0'
            elif col_idx == 8:
                # Value is decimal (e.g. -0.09), format as % multiplies by 100 → -9.00%
                cell.number_format = '0.00%;-0.00%;"-"'
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color="34D399" if is_up else "F87171")
            if col_idx == 1:
                cell.font = Font(name="Arial", size=10, bold=True, color="F1F5F9")

    widths = [10, 14, 12, 12, 12, 12, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    note_row = len(all_data) + 4
    ws.cell(row=note_row, column=1,
            value=f"Source: Polygon.io /v2/aggs · Adjusted · Pulled {date.today()}")
    ws.cell(row=note_row, column=1).font = Font(name="Arial", size=9, color="475569", italic=True)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Fetching {len(TICKERS)} tickers from Polygon.io...")
    print(f"Date range: {START_DATE} → {END_DATE}\n")

    all_data = {}
    for i, ticker in enumerate(TICKERS):
        print(f"  Pulling {ticker}...", end=" ")
        bars = fetch_aggs(ticker)
        all_data[ticker] = bars
        print(f"{len(bars)} bars" if bars else "no data")
        if i < len(TICKERS) - 1:
            time.sleep(13)  # Polygon free tier = 5 calls/min → 1 call per 12s

    print(f"\nBuilding Excel workbook...")
    wb = openpyxl.Workbook()

    # Summary sheet first
    build_summary_sheet(wb, all_data)

    # One sheet per ticker
    for ticker, bars in all_data.items():
        if bars:
            build_ticker_sheet(wb, ticker, bars)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved → {OUTPUT_FILE}")
    print(f"   Sheets: Summary + {sum(1 for b in all_data.values() if b)} ticker tabs")


if __name__ == "__main__":
    main()
