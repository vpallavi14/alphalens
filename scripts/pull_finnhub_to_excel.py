"""
Pull Finnhub company news for multiple tickers and save to Excel.
NOTE: /news-sentiment requires Finnhub Premium. This script uses the FREE
      /company-news endpoint and scores headlines with simple keyword sentiment.
      In Week 3 we'll replace keyword scoring with FinBERT (ML model).

SETUP FIRST:
  1. Sign up free at https://finnhub.io
  2. Copy your API key from the dashboard
  3. Add to your .env file: FINNHUB_API_KEY=your_key_here

Run: python scripts/pull_finnhub_to_excel.py
Output: scripts/output/finnhub_sentiment_data.xlsx
"""

import os, requests, time
from datetime import date, datetime, timedelta
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

API_KEY = os.environ.get("FINNHUB_API_KEY", "")  # Add FINNHUB_API_KEY to your .env
BASE_URL = "https://finnhub.io/api/v1"

TICKERS = ["AAPL", "MSFT", "NVDA", "TSLA", "GOOGL", "AMZN", "META", "NFLX", "AMD", "INTC"]
NEWS_DAYS_BACK = 7  # how many days of news to fetch

OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "finnhub_sentiment_data.xlsx")

# ── Colours ──────────────────────────────────────────────────
HEADER_BG = "0F172A"
HEADER_FG = "F1F5F9"
ALT_ROW   = "1E293B"

def thin_border():
    s = Side(style="thin", color="334155")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell):
    cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
    cell.fill      = PatternFill("solid", start_color=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()

def sentiment_color(score: float) -> str:
    """Map -1..+1 score to a hex color."""
    if score >= 0.2:   return "34D399"  # green
    if score <= -0.2:  return "F87171"  # red
    return "FCD34D"                     # yellow = neutral

def score_to_label(score: float) -> str:
    if score >= 0.5:   return "Very Positive"
    if score >= 0.2:   return "Positive"
    if score >= -0.2:  return "Neutral"
    if score >= -0.5:  return "Negative"
    return "Very Negative"

def score_to_1_10(score: float) -> float:
    """Map -1..+1 sentiment score to 1-10 scale."""
    return round((score + 1) / 2 * 9 + 1, 1)

# ─────────────────────────────────────────────────────────────
# Finnhub API calls
# ─────────────────────────────────────────────────────────────

def fetch_company_news(ticker: str) -> list:
    """Get recent news articles for a ticker."""
    end   = date.today()
    start = end - timedelta(days=NEWS_DAYS_BACK)
    r = requests.get(f"{BASE_URL}/company-news", params={
        "symbol": ticker, "from": str(start), "to": str(end), "token": API_KEY
    }, timeout=15)
    return r.json() if r.status_code == 200 else []

POSITIVE_WORDS = {"beats", "record", "surge", "growth", "profit", "upgrade",
                  "buy", "bullish", "rally", "rises", "gain", "strong", "outperform",
                  "raises", "exceeds", "soars", "momentum", "breakthrough", "wins"}
NEGATIVE_WORDS = {"miss", "drop", "fall", "loss", "downgrade", "sell", "bearish",
                  "decline", "cut", "layoff", "warn", "recall", "investigation",
                  "lawsuit", "fraud", "crash", "disappoints", "weak", "risk", "concern"}

def keyword_sentiment(articles: list) -> dict:
    """
    Simple keyword-based sentiment scoring on headlines.
    Returns scores in the same format as the Finnhub premium endpoint.
    Week 3 will replace this with FinBERT for true NLP sentiment.
    """
    if not articles:
        return {"sentiment": {"score": 0, "bullishPercent": 0, "bearishPercent": 0},
                "buzz": {"articlesInLastWeek": 0, "buzz": 0, "weeklyAverage": 0}}

    pos, neg, neutral = 0, 0, 0
    for article in articles:
        words = set(article.get("headline", "").lower().split())
        has_pos = bool(words & POSITIVE_WORDS)
        has_neg = bool(words & NEGATIVE_WORDS)
        if has_pos and not has_neg:
            pos += 1
        elif has_neg and not has_pos:
            neg += 1
        else:
            neutral += 1

    total = len(articles)
    score = (pos - neg) / total  # -1 to +1
    return {
        "sentiment": {
            "score": round(score, 3),
            "bullishPercent": round(pos / total, 3),
            "bearishPercent": round(neg / total, 3),
        },
        "buzz": {
            "articlesInLastWeek": total,
            "buzz": round(total / 10, 2),   # normalised buzz proxy
            "weeklyAverage": total,
        }
    }

def fetch_quote(ticker: str) -> dict:
    """Get latest price quote."""
    r = requests.get(f"{BASE_URL}/quote", params={
        "symbol": ticker, "token": API_KEY
    }, timeout=15)
    return r.json() if r.status_code == 200 else {}

# ─────────────────────────────────────────────────────────────
# Excel builders
# ─────────────────────────────────────────────────────────────

def build_summary_sheet(wb, summary_rows: list):
    ws = wb.active
    ws.title = "Sentiment Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = f"AlphaLens — Finnhub Sentiment Summary  |  Pulled {date.today()}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="10B981")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = [
        "Ticker", "Last Price", "Bullish %", "Bearish %",
        "Sentiment Score\n(-1 to +1)", "Social Score\n(1–10)",
        "Label", "Articles\n(7d)", "Buzz Score", "News Vol"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        style_header(cell)
    ws.row_dimensions[2].height = 30

    for row_idx, row in enumerate(summary_rows, 3):
        score  = row.get("score", 0)
        alt    = PatternFill("solid", start_color=ALT_ROW if row_idx % 2 == 0 else "0F172A")

        vals = [
            row.get("ticker"),
            row.get("last_price"),
            row.get("bullish_pct"),
            row.get("bearish_pct"),
            score,
            score_to_1_10(score),
            score_to_label(score),
            row.get("article_count"),
            row.get("buzz_score"),
            row.get("weekly_avg"),
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10, color="CBD5E1")
            cell.fill      = alt
            cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "center")
            cell.border    = thin_border()

            if col_idx == 1:
                cell.font = Font(name="Arial", size=10, bold=True, color="F1F5F9")
            elif col_idx == 2:
                cell.number_format = "$#,##0.00"
            elif col_idx in (3, 4):
                cell.number_format = "0.0%"
            elif col_idx == 5:
                cell.number_format = "0.00"
                cell.font = Font(name="Arial", size=10, bold=True, color=sentiment_color(score))
            elif col_idx == 6:
                cell.font = Font(name="Arial", size=10, bold=True, color=sentiment_color(score))
            elif col_idx == 7:
                cell.font = Font(name="Arial", size=10, color=sentiment_color(score))
                cell.alignment = Alignment(horizontal="center")

    widths = [10, 12, 12, 12, 16, 14, 16, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_news_sheet(wb, ticker: str, articles: list, sentiment: dict):
    ws = wb.create_sheet(ticker)
    ws.sheet_view.showGridLines = False

    # Sentiment summary at top
    score = sentiment.get("sentiment", {}).get("score", 0)
    social_score = score_to_1_10(score)

    ws.merge_cells("A1:G1")
    ws["A1"] = f"{ticker} — News & Sentiment  |  Social Score: {social_score}/10  |  {score_to_label(score)}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=12, color=sentiment_color(score))
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Sentiment stats row
    sent = sentiment.get("sentiment", {})
    buzz = sentiment.get("buzz", {})
    ws["A2"] = f"Bullish: {sent.get('bullishPercent', 0)*100:.1f}%"
    ws["B2"] = f"Bearish: {sent.get('bearishPercent', 0)*100:.1f}%"
    ws["C2"] = f"Score: {sent.get('score', 0):.2f}"
    ws["D2"] = f"Articles (7d): {buzz.get('articlesInLastWeek', 0)}"
    ws["E2"] = f"Buzz: {buzz.get('buzz', 0):.2f}"
    for col in range(1, 6):
        ws.cell(row=2, column=col).font = Font(name="Arial", size=10, color="94A3B8")
    ws.row_dimensions[2].height = 18

    # News headers
    headers = ["Unix Timestamp", "Published (UTC)", "Published (ET)", "Headline", "Source", "Category", "URL"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 18

    # Articles
    for row_idx, article in enumerate(articles[:30], 4):  # cap at 30 articles
        unix_ts  = article.get("datetime", 0)
        dt_utc   = datetime.utcfromtimestamp(unix_ts)
        # ET = UTC-4 (EDT) or UTC-5 (EST) — approximate with UTC-4 for summer
        dt_et    = datetime.utcfromtimestamp(unix_ts - 4*3600)
        alt = PatternFill("solid", start_color=ALT_ROW if row_idx % 2 == 0 else "0F172A")

        row_vals = [
            unix_ts,
            dt_utc.strftime("%Y-%m-%d %H:%M:%S") + " UTC",
            dt_et.strftime("%Y-%m-%d %H:%M:%S") + " ET",
            article.get("headline", ""),
            article.get("source", ""),
            article.get("category", ""),
            article.get("url", ""),
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = alt
            cell.border = thin_border()
            if col_idx == 1:  # Unix ms
                cell.font      = Font(name="Arial", size=9, color="475569")
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            elif col_idx in (2, 3):  # timestamps
                cell.font      = Font(name="Arial", size=10, color="7DD3FC")
                cell.alignment = Alignment(horizontal="left")
            elif col_idx == 4:  # headline
                cell.font      = Font(name="Arial", size=10, color="F1F5F9")
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            else:
                cell.font      = Font(name="Arial", size=10, color="94A3B8")
                cell.alignment = Alignment(horizontal="left")

    ws.column_dimensions["A"].width = 16   # Unix timestamp
    ws.column_dimensions["B"].width = 24   # Published UTC
    ws.column_dimensions["C"].width = 24   # Published ET
    ws.column_dimensions["D"].width = 55   # Headline
    ws.column_dimensions["E"].width = 16   # Source
    ws.column_dimensions["F"].width = 14   # Category
    ws.column_dimensions["G"].width = 50   # URL

# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────

def main():
    if not API_KEY:
        print("❌ FINNHUB_API_KEY not set. Add it to your .env file.")
        print("   Sign up free at https://finnhub.io to get your key.")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Fetching sentiment data for {len(TICKERS)} tickers from Finnhub...\n")

    wb = openpyxl.Workbook()
    summary_rows = []

    for ticker in TICKERS:
        print(f"  {ticker}...", end=" ")
        articles  = fetch_company_news(ticker)
        quote     = fetch_quote(ticker)
        sentiment = keyword_sentiment(articles)   # free keyword scoring (FinBERT in Week 3)

        sent  = sentiment.get("sentiment", {})
        buzz  = sentiment.get("buzz", {})
        score = sent.get("score", 0)

        summary_rows.append({
            "ticker":       ticker,
            "last_price":   quote.get("c", 0),
            "bullish_pct":  sent.get("bullishPercent", 0),
            "bearish_pct":  sent.get("bearishPercent", 0),
            "score":        score,
            "article_count": buzz.get("articlesInLastWeek", 0),
            "buzz_score":   buzz.get("buzz", 0),
            "weekly_avg":   buzz.get("weeklyAverage", 0),
        })

        if articles or sentiment:
            build_news_sheet(wb, ticker, articles, sentiment)
            print(f"{len(articles)} articles · score={score:.2f}")
        else:
            print("no data")

        time.sleep(0.5)   # respect free tier rate limit (60 calls/min)

    build_summary_sheet(wb, summary_rows)
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved → {OUTPUT_FILE}")
    print(f"   Sheets: Sentiment Summary + {len(TICKERS)} ticker tabs")

if __name__ == "__main__":
    main()
